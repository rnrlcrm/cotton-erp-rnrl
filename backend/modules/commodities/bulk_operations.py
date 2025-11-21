"""
Bulk Operations for Commodities

Handles Excel import/export and bulk operations.
"""

from __future__ import annotations

import io
from typing import Any, Dict, List
from uuid import UUID

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.ext.asyncio import AsyncSession

from backend.modules.commodities.models import Commodity, CommodityVariety
from backend.modules.commodities.repositories import (
    CommodityRepository,
    CommodityVarietyRepository,
)
from backend.modules.commodities.schemas import CommodityCreate, CommodityVarietyCreate


class BulkOperationService:
    """Service for bulk import/export operations"""
    
    def __init__(self, db: AsyncSession, current_user_id: UUID):
        self.db = db
        self.current_user_id = current_user_id
        self.commodity_repo = CommodityRepository(db)
        self.variety_repo = CommodityVarietyRepository(db)
    
    async def export_commodities_to_excel(
        self,
        commodities: List[Commodity],
        include_varieties: bool = False,
    ) -> bytes:
        """
        Export commodities to Excel file.
        
        Returns Excel file as bytes.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Commodities"
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Commodity sheet headers
        headers = [
            "ID", "Name", "Category", "HSN Code", "GST Rate (%)",
            "Description", "UOM", "Active", "Created At"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        # Data rows
        for row_num, commodity in enumerate(commodities, 2):
            ws.cell(row=row_num, column=1, value=str(commodity.id))
            ws.cell(row=row_num, column=2, value=commodity.name)
            ws.cell(row=row_num, column=3, value=commodity.category)
            ws.cell(row=row_num, column=4, value=commodity.hsn_code)
            ws.cell(row=row_num, column=5, value=float(commodity.gst_rate) if commodity.gst_rate else None)
            ws.cell(row=row_num, column=6, value=commodity.description)
            ws.cell(row=row_num, column=7, value=commodity.uom)
            ws.cell(row=row_num, column=8, value="Yes" if commodity.is_active else "No")
            ws.cell(row=row_num, column=9, value=commodity.created_at.isoformat() if commodity.created_at else None)
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 50)
        
        # Add varieties sheet if requested
        if include_varieties:
            ws_varieties = wb.create_sheet("Varieties")
            variety_headers = [
                "ID", "Commodity ID", "Commodity Name", "Variety Name",
                "Code", "Description", "Is Standard", "Active"
            ]
            
            for col_num, header in enumerate(variety_headers, 1):
                cell = ws_varieties.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
            
            row_num = 2
            for commodity in commodities:
                if commodity.varieties:
                    for variety in commodity.varieties:
                        ws_varieties.cell(row=row_num, column=1, value=str(variety.id))
                        ws_varieties.cell(row=row_num, column=2, value=str(commodity.id))
                        ws_varieties.cell(row=row_num, column=3, value=commodity.name)
                        ws_varieties.cell(row=row_num, column=4, value=variety.name)
                        ws_varieties.cell(row=row_num, column=5, value=variety.code)
                        ws_varieties.cell(row=row_num, column=6, value=variety.description)
                        ws_varieties.cell(row=row_num, column=7, value="Yes" if variety.is_standard else "No")
                        ws_varieties.cell(row=row_num, column=8, value="Yes" if variety.is_active else "No")
                        row_num += 1
            
            # Adjust column widths
            for col in ws_varieties.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws_varieties.column_dimensions[column].width = min(max_length + 2, 50)
        
        # Save to bytes
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        return excel_file.read()
    
    async def import_commodities_from_excel(
        self,
        file_content: bytes,
        skip_errors: bool = True,
    ) -> Dict[str, Any]:
        """
        Import commodities from Excel file.
        
        Returns:
            {
                "success": 10,
                "failed": 2,
                "errors": [...],
                "created_ids": [...]
            }
        """
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        ws = wb.active
        
        results = {
            "success": 0,
            "failed": 0,
            "errors": [],
            "created_ids": [],
        }
        
        # Skip header row
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row[1]:  # Skip if name is empty
                continue
            
            try:
                # Parse row data
                commodity_data = CommodityCreate(
                    name=row[1],
                    category=row[2] or "",
                    hsn_code=row[3],
                    gst_rate=float(row[4]) if row[4] else None,
                    description=row[5],
                    uom=row[6],
                    is_active=str(row[7]).lower() in ("yes", "true", "1") if row[7] else True,
                )
                
                # Create commodity (using repository directly for bulk)
                commodity = await self.commodity_repo.create(**commodity_data.model_dump())
                await self.db.flush()
                
                results["success"] += 1
                results["created_ids"].append(str(commodity.id))
                
            except Exception as e:
                results["failed"] += 1
                results["errors"].append({
                    "row": row_num,
                    "error": str(e),
                    "data": row[1] if len(row) > 1 else None
                })
                
                if not skip_errors:
                    await self.db.rollback()
                    raise
        
        # Commit all successful imports
        if results["success"] > 0:
            await self.db.commit()
        
        return results
    
    async def get_excel_template(self) -> bytes:
        """
        Generate Excel template for commodity import.
        
        Returns template with headers and example data.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Commodities Template"
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        headers = [
            "ID (Leave blank for new)", "Name *", "Category *", "HSN Code",
            "GST Rate (%)", "Description", "UOM", "Active (Yes/No)"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        # Example data
        example_data = [
            ["", "Raw Cotton", "Natural Fiber", "5201", "5.0", "High quality raw cotton", "MT", "Yes"],
            ["", "Cotton Yarn", "Textile", "5205", "12.0", "Combed cotton yarn", "Bales", "Yes"],
        ]
        
        for row_num, data in enumerate(example_data, 2):
            for col_num, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 50)
        
        # Instructions sheet
        ws_instructions = wb.create_sheet("Instructions")
        instructions = [
            ["Commodity Import Template - Instructions"],
            [""],
            ["Required Fields (marked with *):"],
            ["- Name: Commodity name (max 200 characters)"],
            ["- Category: Commodity category (e.g., Natural Fiber, Synthetic, Blended)"],
            [""],
            ["Optional Fields:"],
            ["- HSN Code: Harmonized System Nomenclature code"],
            ["- GST Rate: Tax rate in percentage (e.g., 5.0 for 5%)"],
            ["- Description: Detailed description"],
            ["- UOM: Unit of Measurement (e.g., MT, Quintals, Bales)"],
            ["- Active: Yes/No (default: Yes)"],
            [""],
            ["Notes:"],
            ["1. Leave ID column blank for new commodities"],
            ["2. Delete example rows before importing your data"],
            ["3. Save as .xlsx format"],
            ["4. Maximum 1000 rows per import"],
        ]
        
        for row_num, instruction in enumerate(instructions, 1):
            ws_instructions.cell(row=row_num, column=1, value=instruction[0])
        
        # Save to bytes
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        return excel_file.read()
